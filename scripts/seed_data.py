"""
Seed script: imports Villa Paris event data from Excel into the database.
Run with: cd /app && python3 scripts/seed_data.py
"""
import sqlite3
import json
from datetime import datetime

DB_PATH = '/app/prisma/prisma/dev.db'

def parse_date(s):
    """Parse DD.MM.YYYY or DD.MM.YY -> YYYY-MM-DD ISO string"""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ('%d.%m.%Y', '%d.%m.%y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    # Try partial dates like "07.2027" or "09.2027"
    parts = s.split('.')
    if len(parts) == 2:
        try:
            m, y = int(parts[0]), int(parts[1])
            if m > 12:
                m, y = y, m
            if y < 100:
                y += 2000
            return f'{y:04d}-{m:02d}-01'
        except ValueError:
            pass
    return None

def date_to_epoch_ms(date_str):
    """Convert YYYY-MM-DD to epoch milliseconds"""
    if not date_str:
        return None
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return int(dt.timestamp() * 1000)
    except:
        return None
    """Extract first integer from pax string like '120+', '90-120', '40-50'"""
    if not s:
        return None
    s = str(s).strip().replace('+', '')
    if '-' in s:
        s = s.split('-')[0]
    try:
        return int(float(s))
    except ValueError:
        return None

def parse_prezzo(s):
    """Extract price float from '130€', '70€', 'PB', etc."""
    if not s or not isinstance(s, str):
        return None
    s = s.strip().replace('€', '').replace(',', '.')
    if s == 'PB' or not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None

def determine_tipo(dettagli):
    """Determine event type from details"""
    if not dettagli:
        return 'Festa Privata/Aziendale'
    d = dettagli.lower()
    if 'matrimonio' in d:
        return 'Matrimonio'
    if 'c18' in d or 'compleanno 18' in d:
        return 'Compleanno 18 Anni'
    if 'comunione' in d or 'cresima' in d:
        return 'Comunione'
    if 'battesimo' in d:
        return 'Battesimo'
    if 'compleanno' in d:
        return 'Compleanno'
    if 'concerto' in d or 'associazione' in d or 'manifestazione' in d:
        return 'Evento Culturale'
    if 'aziendale' in d or 'pranzo aziendale' in d:
        return 'Festa Privata/Aziendale'
    return 'Festa Privata/Aziendale'

def determine_stato(sezione):
    """Map section to event status"""
    s = sezione.lower() if sezione else ''
    if 'confermati' in s or 'confermato' in s:
        return 'confermato'
    return 'in_attesa'

def determine_fascia(dettagli):
    """Determine pranzo/cena from details"""
    if not dettagli:
        return 'pranzo'
    d = dettagli.lower()
    if 'cena' in d:
        return 'cena'
    return 'pranzo'

def extract_names_from_details(dettagli, nomi):
    """Extract person names from details or nomi field"""
    if nomi and isinstance(nomi, str) and nomi.strip():
        return nomi.strip()
    if not dettagli:
        return ''
    return dettagli.split(';')[0].strip()

def build_titolo(tipo, dettagli, nomi):
    """Build a descriptive event title"""
    name = extract_names_from_details(dettagli, nomi)
    if tipo == 'Matrimonio':
        # Try to extract couple names
        if name and 'matrimonio' not in name.lower():
            return f'Matrimonio {name}'
        # Parse from details
        parts = (dettagli or '').split(';')
        for p in parts:
            p = p.strip()
            if p and 'matrimonio' not in p.lower() and not any(kw in p.lower() for kw in ['piano', 'carne', 'pesce', 'pranzo', 'cena', 'rito', 'pax', 'salone']):
                return f'Matrimonio {p}'
        return 'Matrimonio'
    return name or tipo

def seed():
    import openpyxl
    wb = openpyxl.load_workbook('/tmp/villa_paris_data.xlsx')
    ws = wb['Tabella_Unica']
    
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    # Clear existing data
    for table in ['EventoCliente', 'VersioneEvento', 'OverrideLog', 'Evento', 'Cliente', 'MenuBase']:
        cur.execute(f'DELETE FROM {table}')
    conn.commit()
    
    # Reset auto-increment
    cur.execute("DELETE FROM sqlite_sequence")
    conn.commit()
    
    now_ms = int(datetime.now().timestamp() * 1000)
    events_created = 0
    clients_created = 0
    
    for row_idx in range(4, 83):  # rows 4-82 are data
        row = {}
        for col in range(1, 22):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(col)
                row[col_letter] = val
        
        sezione = str(row.get('A', '')).strip()
        if not sezione or sezione in ('Sintesi settimana', 'Note organizzative', 'Fonte', 'Documento'):
            continue
        
        data_evento_raw = str(row.get('D', '')).strip() if row.get('D') else ''
        pax_raw = row.get('K', '')
        dettagli = str(row.get('I', '')).strip() if row.get('I') else ''
        nomi = str(row.get('F', '')).strip() if row.get('F') else ''
        contatti = str(row.get('G', '')).strip() if row.get('G') else ''
        canale = str(row.get('H', '')).strip() if row.get('H') else ''
        prezzo_raw = str(row.get('L', '')).strip() if row.get('L') else ''
        localita = str(row.get('S', '')).strip() if row.get('S') else ''
        email = str(row.get('Q', '')).strip() if row.get('Q') else ''
        telefono_raw = str(row.get('R', '')).strip() if row.get('R') else ''
        data_record = str(row.get('C', '')).strip() if row.get('C') else ''
        
        # Skip empty date rows
        data_evento = parse_date(data_evento_raw)
        if not data_evento and not dettagli:
            continue
        
        tipo = determine_tipo(dettagli)
        stato = determine_stato(sezione)
        fascia = determine_fascia(dettagli)
        pax = parse_pax(pax_raw)
        prezzo = parse_prezzo(prezzo_raw)
        titolo = build_titolo(tipo, dettagli, nomi)
        
        # Clean telefono (remove dates that got mixed in)
        telefono = ''
        if telefono_raw:
            # Sometimes dates get mixed into telefono column - filter them
            parts = telefono_raw.split(';')
            for p in parts:
                p = p.strip()
                if p and (p.startswith('3') or p.startswith('+3')) and len(p) >= 9:
                    telefono = p
                    break
        
        # Extract contact info from contatti column if not in separate fields
        if not email and contatti:
            parts = contatti.split(';')
            for p in parts:
                p = p.strip()
                if '@' in p:
                    email = p
                elif (p.startswith('3') or p.startswith('+3')) and len(p) >= 9:
                    telefono = telefono or p
        
        # Map canale to our system
        canale_mapped = ''
        if canale:
            cl = canale.lower()
            if 'matrimonio.com' in cl:
                canale_mapped = 'matrimonio.com'
            elif 'e-mail' in cl or 'email' in cl:
                canale_mapped = 'email'
            elif 'telefon' in cl:
                canale_mapped = 'telefono'
            elif 'social' in cl:
                canale_mapped = 'social'
        
        # Create client
        client_name = extract_names_from_details(dettagli, nomi)
        if not client_name or client_name.lower() in ('matrimonio', ''):
            # Try to extract from dettagli
            parts = (dettagli or '').split(';')
            for p in parts:
                p = p.strip()
                if p and not any(kw in p.lower() for kw in ['matrimonio', 'piano', 'carne', 'pesce', 'pranzo', 'cena', 'rito', 'pax', 'salone', 'giardino', 'c18', 'comunione', 'concerto', 'sta cercando', 'invitati', 'conferma']):
                    client_name = p
                    break
        
        if not client_name:
            client_name = titolo
        
        # Parse name parts
        name_parts = client_name.split(' e ')
        sposa_name = name_parts[0].strip() if name_parts else client_name
        sposo_name = name_parts[1].strip() if len(name_parts) > 1 else ''
        
        # Determine client tipo
        if tipo == 'Matrimonio':
            cliente_tipo = 'sposa'
        elif 'C18' in (dettagli or '') or 'Compleanno' in tipo:
            cliente_tipo = 'festeggiato'
        else:
            cliente_tipo = 'altro'
        
        # Generate unique email if none
        if not email:
            safe = sposa_name.lower().replace(' ', '.').replace("'", '')
            email = f'{safe}.{row_idx}@villa-paris.local'
        
        # Create main client
        cur.execute('''INSERT INTO Cliente (nome, cognome, email, telefono, tipoCliente, canalePrimoContatto, citta, dataPrimoContatto, createdAt, updatedAt)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (sposa_name.split()[0] if ' ' in sposa_name else sposa_name,
                     ' '.join(sposa_name.split()[1:]) if ' ' in sposa_name else '',
                     email,
                     telefono or None,
                     cliente_tipo,
                     canale_mapped or None,
                     localita or None,
                     date_to_epoch_ms(parse_date(data_record)) or now_ms,
                     now_ms, now_ms))
        cliente1_id = cur.lastrowid
        clients_created += 1
        
        # Create sposo client if matrimonio with couple
        cliente2_id = None
        if tipo == 'Matrimonio' and sposo_name:
            safe2 = sposo_name.lower().replace(' ', '.').replace("'", '')
            email2 = f'{safe2}.{row_idx}@villa-paris.local'
            cur.execute('''INSERT INTO Cliente (nome, cognome, email, telefono, tipoCliente, canalePrimoContatto, citta, dataPrimoContatto, createdAt, updatedAt)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (sposo_name.split()[0] if ' ' in sposo_name else sposo_name,
                         ' '.join(sposo_name.split()[1:]) if ' ' in sposo_name else '',
                         email2,
                         None,
                         'sposo',
                         canale_mapped or None,
                         localita or None,
                         date_to_epoch_ms(parse_date(data_record)) or now_ms,
                         now_ms, now_ms))
            cliente2_id = cur.lastrowid
            clients_created += 1
        
        # Create event
        date_proposte = json.dumps([data_evento] if data_evento else [])
        data_conf_ms = date_to_epoch_ms(data_evento) if data_evento and stato == 'confermato' else None
        data_primo = parse_date(data_record)
        data_primo_ms = date_to_epoch_ms(data_primo) or now_ms
        
        cur.execute('''INSERT INTO Evento (tipo, titolo, dateProposte, dataConfermata, fascia, personePreviste, note, stato, 
                        menu, struttura, sposa, sposo, dataPrimoContatto, canalePrimoContatto, createdAt, updatedAt)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (tipo, titolo, date_proposte, data_conf_ms, fascia, pax,
                     dettagli or '', stato, '{}', '{}',
                     sposa_name or None,
                     sposo_name or None,
                     data_primo_ms,
                     canale_mapped or None,
                     now_ms, now_ms))
        evento_id = cur.lastrowid
        events_created += 1
        
        # Link clients
        cur.execute('INSERT INTO EventoCliente (eventoId, clienteId) VALUES (?, ?)', (evento_id, cliente1_id))
        if cliente2_id:
            cur.execute('INSERT INTO EventoCliente (eventoId, clienteId) VALUES (?, ?)', (evento_id, cliente2_id))
    
    conn.commit()
    conn.close()
    
    print(f'Seed completato: {events_created} eventi, {clients_created} clienti creati.')

if __name__ == '__main__':
    seed()
